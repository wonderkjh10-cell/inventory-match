# -*- coding: utf-8 -*-
"""
재고/유통기한 매칭 프로그램
입력 (한 번에 다중 선택, 순서 무관):
  1) 재고현황조회Ⅱ_다운로드.xlsx  - 시트명에 '재고현황' 또는 첫셀에 '재고현황조회'
  2) 진벌리 제품 위치(2024).xlsx - '진벌리' 시트 보유
  3) buyList_*.xlsx              - '구매업로드' 시트 보유
처리:
  - 재고현황의 '현재고-가용' (Q열) → 진벌리 E열(재고)
  - buyList의 '소비기한' (Y열, 동일 품번이면 최대값) → 진벌리 G열(유통기한)
출력:
  - 바탕화면\재고매칭\YYYY-MM-DD 오전(또는 오후).xlsx
  - 재고현황 파일은 작업 후 삭제
"""
import sys
import os
import re
import json
import zipfile
import shutil
import webbrowser
import urllib.request
from pathlib import Path
from send2trash import send2trash
from openpyxl import load_workbook
from datetime import datetime, date
import tkinter as tk
from tkinter import filedialog, messagebox

GITHUB_REPO = "wonderkjh10-cell/inventory-match"
UPDATE_CHECK_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"


def _resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def get_current_version() -> str:
    try:
        with open(_resource_path("version.txt"), "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return "0.0.0"


def check_for_update():
    """최신 버전 확인. (latest_version, download_url) 또는 None."""
    try:
        req = urllib.request.Request(
            UPDATE_CHECK_URL,
            headers={"Accept": "application/vnd.github.v3+json"},
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode("utf-8"))
        latest = data.get("tag_name", "").lstrip("v")
        current = get_current_version()
        if latest and latest != current:
            download_url = None
            for asset in data.get("assets", []):
                if asset["name"].endswith(".exe"):
                    download_url = asset["browser_download_url"]
                    break
            return (latest, download_url)
    except Exception:
        pass
    return None


def perform_self_update(download_url: str, latest_version: str) -> None:
    """새 exe를 다운로드한 뒤 배치 파일로 자기 자신을 교체하고 재시작."""
    if not getattr(sys, "frozen", False):
        # 개발 환경(파이썬 직접 실행)에서는 브라우저로 폴백
        webbrowser.open(download_url)
        sys.exit(0)

    current_exe = Path(sys.executable)
    new_exe = current_exe.parent / f"재고매칭_v{latest_version}.exe"
    if new_exe.exists():
        # 같은 버전 파일이 이미 있으면 임시명으로
        new_exe = current_exe.parent / f"재고매칭_v{latest_version}_new.exe"

    # 진행 상황 표시용 작은 창
    progress = tk.Toplevel()
    progress.title("업데이트 다운로드 중")
    progress.geometry("360x80")
    progress.attributes("-topmost", True)
    label = tk.Label(progress, text=f"새 버전 {latest_version} 다운로드 중...\n잠시만 기다려주세요.", padx=20, pady=20)
    label.pack()
    progress.update()

    try:
        urllib.request.urlretrieve(download_url, new_exe)
    except Exception as e:
        progress.destroy()
        messagebox.showerror("업데이트 실패", f"다운로드 중 오류:\n{e}")
        return

    progress.destroy()

    # 배치: 현재 exe 종료 대기 → 새 버전 exe 실행 → 이전 버전 exe 삭제 → 자기 삭제
    bat_path = current_exe.with_name("__update__.bat")
    bat_content = (
        "@echo off\r\n"
        "chcp 65001 >nul\r\n"
        "ping 127.0.0.1 -n 2 >nul\r\n"
        f':loop\r\n'
        f'del "{current_exe}" >nul 2>&1\r\n'
        f'if exist "{current_exe}" (\r\n'
        f'  ping 127.0.0.1 -n 2 >nul\r\n'
        f'  goto loop\r\n'
        f')\r\n'
        f'start "" "{new_exe}"\r\n'
        f'del "%~f0"\r\n'
    )
    bat_path.write_text(bat_content, encoding="utf-8")

    # 백그라운드로 배치 실행 (콘솔창 안 뜸)
    import subprocess
    CREATE_NO_WINDOW = 0x08000000
    DETACHED_PROCESS = 0x00000008
    subprocess.Popen(
        ["cmd", "/c", str(bat_path)],
        creationflags=CREATE_NO_WINDOW | DETACHED_PROCESS,
        close_fds=True,
    )
    sys.exit(0)


def prompt_update_if_available():
    info = check_for_update()
    if info is None:
        return
    latest, url = info
    current = get_current_version()
    if not url:
        return
    answer = messagebox.askyesno(
        "업데이트",
        f"새 버전이 출시되었습니다.\n\n"
        f"  현재 버전: {current}\n"
        f"  최신 버전: {latest}\n\n"
        f"지금 자동으로 업데이트할까요?\n"
        f"(다운로드 후 프로그램이 자동으로 재시작됩니다)",
    )
    if answer:
        perform_self_update(url, latest)


def get_desktop() -> Path:
    userprofile = os.environ.get("USERPROFILE")
    if userprofile:
        for sub in ("Desktop", "OneDrive/Desktop", "OneDrive/바탕 화면"):
            p = Path(userprofile) / sub
            if p.exists():
                return p
    return Path.home() / "Desktop"


def pick_files(title: str) -> list[Path]:
    paths = filedialog.askopenfilenames(
        title=title,
        filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")],
    )
    return [Path(p) for p in paths]


def classify_file(path: Path) -> str:
    """파일을 보고 'stock' / 'location' / 'buylist' / 'unknown' 판별."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        first_cell = None
        try:
            ws = wb.active
            first_cell = ws.cell(row=1, column=1).value
        except Exception:
            pass
        wb.close()
    except Exception:
        return "unknown"

    if "진벌리" in sheet_names:
        return "location"
    if "구매업로드" in sheet_names:
        return "buylist"
    for s in sheet_names:
        if "재고현황" in s or "다운로드" in s:
            return "stock"
    if first_cell and "재고현황조회" in str(first_cell):
        return "stock"
    return "unknown"


def build_stock_map(stock_path: Path) -> dict:
    """재고현황 → {상품코드: 현재고가용}."""
    wb = load_workbook(stock_path, data_only=True)
    ws = wb.active
    CODE_COL = 3
    CURRENT_STOCK_COL = 17

    stock_map = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if len(row) < CURRENT_STOCK_COL:
            continue
        code = row[CODE_COL - 1]
        qty = row[CURRENT_STOCK_COL - 1]
        if code is None:
            continue
        code = str(code).strip()
        if not code:
            continue
        try:
            qty = int(qty) if qty is not None else 0
        except (ValueError, TypeError):
            qty = 0
        stock_map[code] = stock_map.get(code, 0) + qty
    return stock_map


def build_expiry_map(buylist_path: Path) -> dict:
    """buyList → {품목코드(6자리 문자열): 가장 늦은 소비기한 datetime}."""
    wb = load_workbook(buylist_path, data_only=True)
    if "구매업로드" in wb.sheetnames:
        ws = wb["구매업로드"]
    else:
        ws = wb.active

    # 헤더에서 '품목코드'와 '소비기한' 컬럼 자동 탐색
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    code_col = None
    expiry_col = None
    for idx, h in enumerate(header, start=1):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str == "품목코드":
            code_col = idx
        elif h_str in ("소비기한", "유통기한"):
            expiry_col = idx
    if code_col is None or expiry_col is None:
        raise ValueError("buyList 파일에서 '품목코드' 또는 '소비기한' 컬럼을 찾지 못했습니다.")

    expiry_map = {}
    for row_idx in range(2, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=code_col).value
        expiry = ws.cell(row=row_idx, column=expiry_col).value
        if code is None or expiry is None:
            continue
        code = str(code).strip()
        if not code:
            continue
        # 정수형 코드를 문자열로 정규화 (소수점 제거)
        if code.endswith(".0"):
            code = code[:-2]

        # 날짜 정규화
        exp_dt = parse_date(expiry)
        if exp_dt is None:
            continue

        prev = expiry_map.get(code)
        if prev is None or exp_dt > prev:
            expiry_map[code] = exp_dt
    return expiry_map


def parse_date(value):
    """엑셀 셀의 다양한 날짜 형식을 datetime으로 변환."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def update_location_file(
    location_path: Path,
    stock_map: dict,
    expiry_map: dict,
    output_path: Path,
) -> dict:
    wb = load_workbook(location_path)
    if "진벌리" not in wb.sheetnames:
        raise ValueError("제품위치 파일에 '진벌리' 시트가 없습니다.")
    ws = wb["진벌리"]

    wb_values = load_workbook(location_path, data_only=True)
    ws_values = wb_values["진벌리"]

    # 진벌리 시트: A=품번, B=상품명, C=위치, D=상품코드, E=재고, F=삭제확인, G=유통기한
    PUMBUN_COL = 1
    CODE_COL = 4
    STOCK_COL = 5
    EXPIRY_COL = 7

    stock_matched = stock_unmatched = 0
    expiry_matched = expiry_unmatched = 0

    for row_idx in range(2, ws.max_row + 1):
        품번 = ws_values.cell(row=row_idx, column=PUMBUN_COL).value
        품번_str = str(품번).strip() if 품번 is not None else ""
        if 품번_str.endswith(".0"):
            품번_str = 품번_str[:-2]

        # 상품코드 (재고 매칭용)
        code = ws_values.cell(row=row_idx, column=CODE_COL).value
        if code is None or (isinstance(code, str) and code.startswith("=")):
            if 품번_str:
                code = f"{품번_str}-0001"
        code_str = str(code).strip() if code is not None else ""

        # 재고 매칭
        if code_str:
            if code_str in stock_map:
                ws.cell(row=row_idx, column=STOCK_COL).value = stock_map[code_str]
                stock_matched += 1
            else:
                stock_unmatched += 1

        # 유통기한 매칭 (품번 6자리 기준)
        if 품번_str and 품번_str in expiry_map:
            cell = ws.cell(row=row_idx, column=EXPIRY_COL)
            cell.value = expiry_map[품번_str]
            cell.number_format = "yyyy-mm-dd"
            expiry_matched += 1
        elif 품번_str:
            expiry_unmatched += 1

    try:
        wb._external_links = []
    except Exception:
        pass

    wb.save(output_path)
    strip_external_links(output_path)
    return {
        "stock_matched": stock_matched,
        "stock_unmatched": stock_unmatched,
        "expiry_matched": expiry_matched,
        "expiry_unmatched": expiry_unmatched,
    }


def strip_external_links(xlsx_path: Path) -> None:
    """저장된 xlsx 내부에서 외부 데이터 연결 완전 제거."""
    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        external_files = {n for n in names if n.startswith("xl/externalLinks/")}

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                if name in external_files:
                    continue
                data = zin.read(name)

                if name == "xl/workbook.xml":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r"<externalReferences>.*?</externalReferences>",
                        "",
                        text,
                        flags=re.DOTALL,
                    )
                    data = text.encode("utf-8")

                elif name == "xl/_rels/workbook.xml.rels":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r'<Relationship[^>]*Type="[^"]*externalLink[^"]*"[^/]*/>',
                        "",
                        text,
                    )
                    data = text.encode("utf-8")

                elif name == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r'<Override[^>]*PartName="/xl/externalLinks/[^"]*"[^/]*/>',
                        "",
                        text,
                    )
                    data = text.encode("utf-8")

                elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    text = data.decode("utf-8")

                    def strip_external_formula(m):
                        cell_xml = m.group(0)
                        if re.search(r'<f[^>]*>[^<]*\[\d+\][^<]*</f>', cell_xml):
                            cell_xml = re.sub(r'<f[^>]*>[^<]*</f>', "", cell_xml)
                        return cell_xml

                    text = re.sub(
                        r'<c\b[^>]*>.*?</c>',
                        strip_external_formula,
                        text,
                        flags=re.DOTALL,
                    )
                    data = text.encode("utf-8")

                zout.writestr(name, data)

    shutil.move(str(tmp_path), str(xlsx_path))


def make_output_path() -> Path:
    output_dir = get_desktop() / "재고매칭"
    output_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    ampm = "오전" if now.hour < 12 else "오후"
    date_str = now.strftime("%Y-%m-%d")
    output_path = output_dir / f"{date_str} {ampm}.xlsx"

    if output_path.exists():
        n = 2
        while True:
            candidate = output_dir / f"{date_str} {ampm} ({n}).xlsx"
            if not candidate.exists():
                output_path = candidate
                break
            n += 1
    return output_path


def main():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    prompt_update_if_available()

    files = pick_files("엑셀 파일 3개를 한번에 선택하세요 (재고현황, 제품위치, buyList — 순서 무관)")
    if len(files) < 3:
        messagebox.showwarning(
            "취소",
            f"엑셀 파일 3개를 함께 선택해야 합니다. (선택된 개수: {len(files)})",
        )
        return

    stock_file = location_file = buylist_file = None
    for p in files:
        kind = classify_file(p)
        if kind == "stock" and stock_file is None:
            stock_file = p
        elif kind == "location" and location_file is None:
            location_file = p
        elif kind == "buylist" and buylist_file is None:
            buylist_file = p

    missing = []
    if stock_file is None:
        missing.append("재고현황 (시트명 '재고현황' 또는 첫셀 '재고현황조회')")
    if location_file is None:
        missing.append("제품위치 ('진벌리' 시트 포함)")
    if buylist_file is None:
        missing.append("buyList ('구매업로드' 시트 포함)")
    if missing:
        messagebox.showerror(
            "오류",
            "다음 파일을 자동 인식하지 못했습니다:\n- " + "\n- ".join(missing),
        )
        return

    print("=" * 60)
    print(f"  재고현황: {stock_file.name}")
    print(f"  제품위치: {location_file.name}")
    print(f"  buyList:  {buylist_file.name}")
    print()

    print("[1/5] 재고현황 읽는 중...")
    stock_map = build_stock_map(stock_file)
    print(f"      -> {len(stock_map)}개 상품코드 로드")

    print("[2/5] buyList 읽는 중...")
    expiry_map = build_expiry_map(buylist_file)
    print(f"      -> {len(expiry_map)}개 품목코드 로드 (동일코드는 최신 소비기한)")

    output_path = make_output_path()
    print("[3/5] 제품위치 갱신 중...")
    result = update_location_file(location_file, stock_map, expiry_map, output_path)
    print(f"      -> 저장: {output_path}")

    print(f"[4/5] 입력 파일 3개를 휴지통으로 이동...")
    delete_ok = True
    deleted = []
    failed = []
    for f in (stock_file, location_file, buylist_file):
        try:
            send2trash(str(f))
            deleted.append(f.name)
        except Exception as e:
            delete_ok = False
            failed.append(f"{f.name}: {e}")
    for n in deleted:
        print(f"      -> 휴지통 이동: {n}")
    for n in failed:
        print(f"      -> 실패: {n}")

    print("[5/5] 작업 완료")
    print("=" * 60)

    msg = (
        f"[재고]\n"
        f"  매칭 성공: {result['stock_matched']}건\n"
        f"  매칭 실패: {result['stock_unmatched']}건\n\n"
        f"[유통기한]\n"
        f"  매칭 성공: {result['expiry_matched']}건\n"
        f"  매칭 실패: {result['expiry_unmatched']}건\n\n"
        f"저장 위치:\n{output_path}"
    )
    if not delete_ok:
        msg += "\n\n[주의] 일부 입력 파일을 휴지통으로 이동하지 못했습니다."
    messagebox.showinfo("완료", msg)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
        try:
            messagebox.showerror("오류", traceback.format_exc())
        except Exception:
            pass
        input("\n엔터를 누르면 종료됩니다...")
